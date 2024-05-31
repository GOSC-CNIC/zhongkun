import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone as dj_timezone

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from apps.service.models import DataCenter
from apps.app_net_manage.models import OrgVirtualObject
from apps.app_net_ipam.models import (
    ASN, IPv4Range, ipv4_str_to_int, ipv6_str_to_bytes, IPv6Range
)


class Command(BaseCommand):
    help = """
    manage.py iprangeimport --file="/home/x.xlsx" --ipv=6; 
    sheet format: ['开始地址', '结束地址', '掩码', '创建时间', '更新时间', 'AS', '地址类型', '状态', '模型', '关联实例', '备注', '机构']
    """

    STATUS_CODE_MAP = {
        '预留': IPv4Range.Status.RESERVED.value,
        '入库': IPv4Range.Status.WAIT.value,
        '已分配': IPv4Range.Status.ASSIGNED.value,
    }

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', dest='filename', type=str, required=True,
            help='path of file',
        )
        parser.add_argument(
            '--ipv', dest='ip_version', required=True, type=int, choices=[4, 6],
            help='IP version range will import.',
        )

        parser.add_argument(
            # 当命令行有此参数时取值const, 否则取值default
            '--test', default=None, nargs='?', dest='test', const=True,
            help='test.',
        )

    def handle(self, *args, **options):
        is_test = options['test']
        filename = options['filename']
        ip_version = options['ip_version']
        if not Path(filename).exists():
            raise CommandError('not found file')

        ipv4_rows, ipv6_rows = self.parse_sheet(filename=filename)
        v4_len = len(ipv4_rows)
        v6_len = len(ipv6_rows)
        self.stdout.write(self.style.WARNING(
            f'All: {v4_len + v6_len}; ipv4 range rows: {v4_len}; ipv6 range rows: {v6_len};'))
        if is_test:
            self.stdout.write(self.style.NOTICE('In mode Test'))

        self.stdout.write(self.style.ERROR(f'Will import IP v{ip_version} ranges.'))
        if input('Are you sure you want to do this?\n\n' + "Type 'yes' to continue, or 'no' to cancel: ") != 'yes':
            raise CommandError("cancelled.")

        if ip_version == 4:
            self.import_ipv4_range_rows(ipv4_rows=ipv4_rows, is_test=is_test)
        elif ip_version == 6:
            self.import_ipv6_range_rows(ipv6_rows=ipv6_rows, is_test=is_test)
        else:
            self.stdout.write(self.style.NOTICE('Nothing do.'))

    def sheet(self, filename: str):
        wb = load_workbook(filename=filename, read_only=False)
        sheetnames = wb.sheetnames
        if not sheetnames:
            self.stdout.write(self.style.WARNING('空文件，not found sheet'))
            return

        if len(sheetnames) > 1:
            print(sheetnames)
            self.stdout.write(self.style.WARNING(f'found {len(sheetnames)} sheet，must be 1 sheet'))
            return

        sheet = wb.worksheets[0]
        self.stdout.write(self.style.WARNING(f'now sheet: {sheet.title}, {sheet.max_row} rows.'))
        return sheet

    def parse_sheet(self, filename: str):
        sheet = self.sheet(filename=filename)

        row_num = 0
        ipv4_rows = []
        ipv6_rows = []
        for row in sheet.iter_rows(min_row=1, max_row=None, max_col=12, values_only=False):
            row_num += 1
            li = []
            for cell in row:
                val = self.get_cell_value(sheet=sheet, cell=cell)
                li.append(val)

            # 不是ip段数据行，跳过
            if not ('.' in li[0] or ':' in li[0]):
                continue

            ip_version, row_data = self.parse_row(row=li)
            if ip_version == 4:
                ipv4_rows.append(row_data)
            else:
                ipv6_rows.append(row_data)

        return ipv4_rows, ipv6_rows

    def parse_row(self, row: list):
        ip_version = 4
        start_ip = row[0]
        end_ip = row[1]
        if ':' in start_ip and ':' in end_ip:     # ipv6
            ip_version = 6
        elif '.' in start_ip and '.' in end_ip:
            ip_version = 4
        else:
            self.stdout.write(self.style.DANGER(f'ip range invalid "{start_ip} - {end_ip}"'))

        status = row[7]
        if status not in self.STATUS_CODE_MAP:
            self.stdout.write(self.style.DANGER(f'status "{status}" not in {self.STATUS_CODE_MAP.keys()}'))

        return ip_version, {
            'start_ip': row[0],
            'end_ip': row[1],
            'mask_len': int(row[2]),
            'create_time': self.time_str_to_datetime(time_str=row[3]),
            'update_time': self.time_str_to_datetime(time_str=row[4]),
            'asn': int(row[5]),
            'ip_version': int(row[6]),
            'status': row[7],
            'model_type': row[8],
            'virt_obj_name': row[9],
            'remark': row[10] if row[10] else '',
            'org_name': row[11]
        }

    def import_ipv4_range_rows(self, ipv4_rows: list, is_test: bool):
        new_count = 0
        for row_data in ipv4_rows:
            ip_range, created = self.get_or_create_ipv4_range(
                start_ip=row_data['start_ip'],
                end_ip=row_data['end_ip'],
                mask_len=row_data['mask_len'],
                create_time=row_data['create_time'],
                update_time=row_data['update_time'],
                asn=row_data['asn'],
                status=row_data['status'],
                virt_obj_name=row_data['virt_obj_name'],
                org_name=row_data['org_name'],
                remark=row_data['remark'],
                is_test=is_test
            )
            if created:
                new_count += 1

        self.stdout.write(self.style.SUCCESS(f'All {len(ipv4_rows)}, new created ipv4 range {new_count}'))

    def get_or_create_ipv4_range(
            self, start_ip: str, end_ip: str, mask_len: int,
            create_time: datetime, update_time: datetime, asn: int, status: str,
            virt_obj_name: str, org_name: str, remark: str, is_test: bool = False
    ):
        """
        :status: 预留、入库、已分配
        :return: (
            obj: IPv4Range,
            created: bool
        )
        """
        start_int = ipv4_str_to_int(ipv4=start_ip)
        end_int = ipv4_str_to_int(ipv4=end_ip)
        mask_len = int(mask_len)
        asn = int(asn)
        status_code = self.convert_status_code(status=status)

        ip_range = IPv4Range.objects.filter(start_address=start_int, end_address=end_int).first()
        if ip_range:
            return ip_range, False

        if is_test or not virt_obj_name:
            virt_obj = None
        else:
            virt_obj = self.get_or_create_virt_obj(name=virt_obj_name, org_name=org_name)

        if status_code == IPv4Range.Status.ASSIGNED.value:
            assigned_time = update_time
        else:
            assigned_time = None

        asn_obj = self.get_or_create_asn(asn=asn)

        ip_range = IPv4Range(
            name='',
            status=status_code,
            creation_time=create_time,
            update_time=update_time,
            assigned_time=assigned_time,
            asn=asn_obj, org_virt_obj=virt_obj,
            start_address=start_int,
            end_address=end_int,
            mask_len=mask_len,
            admin_remark=remark if remark else ''
        )
        ip_range.name = str(ip_range.start_address_network)
        ip_range.clean()

        if not is_test:
            ip_range.save(force_insert=True)

        return ip_range, True

    def import_ipv6_range_rows(self, ipv6_rows: list, is_test: bool):
        new_count = 0
        for row_data in ipv6_rows:
            ip_range, created = self.get_or_create_ipv6_range(
                start_ip=row_data['start_ip'],
                end_ip=row_data['end_ip'],
                prefixlen=row_data['mask_len'],
                create_time=row_data['create_time'],
                update_time=row_data['update_time'],
                asn=row_data['asn'],
                status=row_data['status'],
                virt_obj_name=row_data['virt_obj_name'],
                org_name=row_data['org_name'],
                remark=row_data['remark'],
                is_test=is_test
            )
            if created:
                new_count += 1

        self.stdout.write(self.style.SUCCESS(f'All {len(ipv6_rows)}, new created ipv6 range {new_count}'))

    def get_or_create_ipv6_range(
            self, start_ip: str, end_ip: str, prefixlen: int,
            create_time: datetime, update_time: datetime, asn: int, status: str,
            virt_obj_name: str, org_name: str, remark: str, is_test: bool = False
    ):
        """
        :status: 预留、入库、已分配
        :return: (
            obj: IPv4Range,
            created: bool
        )
        """
        start_bytes = ipv6_str_to_bytes(ipv6=start_ip)
        end_bytes = ipv6_str_to_bytes(ipv6=end_ip)
        prefixlen = int(prefixlen)
        asn = int(asn)
        status_code = self.convert_status_code(status=status)

        ip_range = IPv6Range.objects.filter(start_address=start_bytes, end_address=end_bytes).first()
        if ip_range:
            return ip_range, False

        if is_test or not virt_obj_name:
            virt_obj = None
        else:
            virt_obj = self.get_or_create_virt_obj(name=virt_obj_name, org_name=org_name)

        if status_code == IPv6Range.Status.ASSIGNED.value:
            assigned_time = update_time
        else:
            assigned_time = None

        asn_obj = self.get_or_create_asn(asn=asn)

        ip_range = IPv6Range(
            name='',
            status=status_code,
            creation_time=create_time,
            update_time=update_time,
            assigned_time=assigned_time,
            asn=asn_obj, org_virt_obj=virt_obj,
            start_address=start_bytes,
            end_address=end_bytes,
            prefixlen=prefixlen,
            admin_remark=remark if remark else ''
        )
        ip_range.name = str(ip_range.start_address_network)
        ip_range.clean()

        if not is_test:
            ip_range.save(force_insert=True)

        return ip_range, True

    @staticmethod
    def time_str_to_datetime(time_str: int, format_='%Y-%m-%d %H:%M:%S'):
        dt = datetime.datetime.strptime(time_str, format_)
        return dt.replace(tzinfo=datetime.timezone(offset=datetime.timedelta(hours=8)))

    @staticmethod
    def get_cell_value(sheet: Worksheet, cell):
        """
        检查是否为合并单元格并获取对应行列单元格的值。
        如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
        :param sheet: 当前工作表对象
        :param cell:
        """
        if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
                if cell.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

        val = cell.value
        if isinstance(val, str):
            val = val.strip(' ')

        return val

    @staticmethod
    def get_or_create_asn(asn: int):
        ins = ASN.objects.filter(number=asn).first()
        if ins is not None:
            return ins

        ins = ASN(number=asn, creation_time=dj_timezone.now())
        ins.save(force_insert=True)
        return ins

    @staticmethod
    def get_or_create_org(name: str):
        org = DataCenter.objects.filter(name=name).first()
        if org is not None:
            return org

        org = DataCenter(name=name, name_en=name, creation_time=dj_timezone.now(), status=DataCenter.STATUS_ENABLE)
        org.save(force_insert=True)
        return org

    def get_or_create_virt_obj(self, name: str, org_name: str):
        ovo = OrgVirtualObject.objects.filter(name=name).first()
        if ovo is not None:
            return ovo

        if not org_name:
            org_name = '其他'

        org = self.get_or_create_org(name=org_name)
        ovo = OrgVirtualObject(name=name, organization=org, creation_time=dj_timezone.now())
        ovo.save(force_insert=True)
        return ovo

    def convert_status_code(self, status: str):
        if status in self.STATUS_CODE_MAP:
            return self.STATUS_CODE_MAP[status]

        return ''
