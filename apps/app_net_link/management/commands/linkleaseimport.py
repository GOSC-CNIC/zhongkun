from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.app_net_link.models import LeaseLine
from apps.app_net_link.managers.link import LeaseLineManager


class Command(BaseCommand):
    help = """
    manage.py linkleaseimport --file="/home/x.csv"; 
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', dest='filename', type=str, required=True,
            help='path of file',
        )

    def handle(self, *args, **options):
        if LeaseLine.objects.all().exists():
            raise CommandError('the leaseline table must be empty')

        filename = options['filename']
        if not Path(filename).exists():
            raise CommandError('not found file')

        leaselines = self.parse_sheet(filename=filename)

        self.stdout.write(self.style.NOTICE(f'Will import leaseline data'))
        self.stdout.write(self.style.NOTICE(f'First leaseline:\n{leaselines[0]}'))
        self.stdout.write(self.style.NOTICE(f'Last leaseline:\n{leaselines[-1]}'))
        self.stdout.write(self.style.NOTICE(f'Total number of leaseline records is {len(leaselines)}'))
        if input('Are you sure you want to do this?\n\n' + "Type 'yes' to continue, or 'no' to cancel: ") != 'yes':
            raise CommandError("cancelled.")
        
        self.do_creat_leaseline(leaselines=leaselines)
        self.stdout.write(self.style.SUCCESS(f'new created leaseline records {len(leaselines)}'))

    def sheet(self, filename: str):
        wb = load_workbook(filename=filename, read_only=False)
        sheetnames = wb.sheetnames
        if not sheetnames:
            self.stdout.write(self.style.WARNING('空文件，not found sheet'))
            return

        if len(sheetnames) > 1:
            print(sheetnames)
            self.stdout.write(self.style.WARNING(
                f'found {len(sheetnames)} sheet，must be 1 sheet'))
            return

        sheet = wb.worksheets[0]
        self.stdout.write(self.style.WARNING(
            f'now sheet: {sheet.title}, {sheet.max_row} rows.'))
        return sheet

    def parse_sheet(self, filename: str):
        sheet = self.sheet(filename=filename)
        row_num = 0
        leaselines = []
        for row in sheet.iter_rows(min_row=2, max_row=None, max_col=14, values_only=False):
            row_num += 1
            row_data = [cell.value for cell in row]
            leaselines.append(
                {
                    'private_line_number': row_data[0] if row_data[0] is not None else '',
                    'lease_line_code': row_data[1] if row_data[1] is not None else '',
                    'line_username': row_data[2] if row_data[2] is not None else '',
                    'endpoint_a': row_data[3] if row_data[3] is not None else '',
                    'endpoint_z': row_data[4] if row_data[4] is not None else '',
                    'line_type': row_data[5] if row_data[5] is not None else '',
                    'cable_type': row_data[6] if row_data[6] is not None else '',
                    'bandwidth': row_data[7] if row_data[7] is not None else None,
                    'length': row_data[8] if row_data[8] is not None else None,
                    'provider': row_data[9] if row_data[9] is not None else '',
                    'enable_date': row_data[10] if row_data[10] is not None else None,
                    'is_whithdrawal': row_data[11],
                    'money': row_data[12] if row_data[12] is not None else None,
                    'remarks': row_data[13] if row_data[13] is not None else '',
                }
            )
        return leaselines

    @staticmethod
    def do_creat_leaseline(leaselines: list):
        for data in leaselines:
            LeaseLineManager.create_leaseline(
                private_line_number=data['private_line_number'],
                lease_line_code=data['lease_line_code'],
                line_username=data['line_username'],
                endpoint_a=data['endpoint_a'],
                endpoint_z=data['endpoint_z'],
                line_type=data['line_type'],
                cable_type=data['cable_type'],
                bandwidth=data['bandwidth'],
                length=data['length'],
                provider=data['provider'],
                enable_date=data['enable_date'],
                is_whithdrawal=data['is_whithdrawal'],
                money=data['money'],
                remarks=data['remarks']
            )
