from openpyxl import load_workbook
import os
import sys

import random
import string
from django import setup
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError

# 将项目路径添加到系统搜寻路径当中，查找方式为从当前脚本开始，找到要调用的django项目的路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'django_site.settings')
setup()

from apps.app_users.models import UserProfile


class Command(BaseCommand):
    help = """
    manage.py bulk_create_user --file="/home/x.xlsx"; 
    Title format: ['通行证账号', '姓', '名']
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', dest='filename', type=str, required=True,
            help='path of file',
        )

        parser.add_argument(
            # 当命令行有此参数时取值const, 否则取值default
            '--test', default=None, nargs='?', dest='test', const=True,
            help='test.',
        )

    def handle(self, *args, **options):
        is_test = options['test']
        filename = options['filename']
        if not Path(filename).exists():
            raise CommandError('not found file')

        user_info_list = self.parser_user_info(filename=filename)
        user_info_len = len(user_info_list)

        self.stdout.write(self.style.WARNING(f'excel rows: {user_info_len};'))
        if is_test:
            self.stdout.write(self.style.NOTICE('In mode Test'))
        if user_info_len:
            new_created = 0
            existed = 0
            for user_info in user_info_list:
                created = self.get_or_create_user(
                    username=user_info[0],
                    last_name=user_info[1],     # 姓
                    first_name=user_info[2]
                )
                if created:
                    new_created += 1
                else:
                    existed += 1
            self.stdout.write(self.style.SUCCESS(f'All {user_info_len}, new created {new_created}, existed {existed};'))

        else:
            self.stdout.write(self.style.NOTICE('Nothing do.'))

    @staticmethod
    def parser_user_info(filename):
        wb = load_workbook(filename=filename, read_only=False)
        sheet = wb.active
        items = []
        for r in sheet.iter_rows(values_only=True):
            item = list(r)
            item = item[:3]
            if item[0] == '通行证账号':
                continue
            items.append(item)
        return items

    @staticmethod
    def generate_random_password(length):
        chars = string.ascii_letters + string.digits
        password = ''.join(random.choice(chars) for _ in range(length))
        return password

    def get_or_create_user(self, username=None, company='', is_superuser=False, first_name='',
                           last_name=''):
        # if not password:
        #     password = self.generate_random_password(8)
        if UserProfile.objects.filter(username=username).first():
            return False

        _, created_status = UserProfile.objects.get_or_create(
            username=username,

            company=company,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
            is_superuser=is_superuser)
        return created_status
