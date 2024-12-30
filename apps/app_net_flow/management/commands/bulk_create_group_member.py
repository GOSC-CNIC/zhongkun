from openpyxl import load_workbook
import os
import sys

import random
import string
from django import setup
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError

# 将项目路径添加到系统搜寻路径当中，查找方式为从当前脚本开始，找到要调用的django项目的路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'django_site.settings')
setup()

from apps.app_users.models import UserProfile
from apps.app_net_flow.models import Menu2Member
from apps.app_net_flow.models import MenuModel


class Command(BaseCommand):
    help = """
    manage.py bulk_create_group_member --file="/home/x.xlsx"; 
    Title format: ['顶级id', '顶级名称', '一级id','一级名称','二级id','二级名称','用户邮箱账号']
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', dest='filename', type=str, required=True,
            help='path of file',
        )

        parser.add_argument(
            # 当命令行有此参数时取值const, 否则取值default
            '--test', default=None, nargs='?', dest='test', const=True,
            help='test.',
        )

    def handle(self, *args, **options):
        is_test = options['test']
        filename = options['filename']
        if not Path(filename).exists():
            raise CommandError('not found file')

        group_member_info_list = self.parser_group_member_info(filename=filename)

        self.stdout.write(self.style.WARNING(f'excel rows: {len(group_member_info_list)};'))
        if is_test:
            self.stdout.write(self.style.NOTICE('In mode Test'))

        if group_member_info_list:
            new_created = 0
            existed = 0
            for member_info in group_member_info_list:
                created = self.create_group_member(
                    group_id=member_info[0],
                    username=member_info[1],
                    role=member_info[2]
                )
                if created:
                    new_created += 1
                else:
                    existed += 1
            self.stdout.write(self.style.SUCCESS(
                f'All {len(group_member_info_list)}, new created {new_created}, existed {existed};')
            )
        else:
            self.stdout.write(self.style.NOTICE('Nothing do.'))

    @staticmethod
    def parser_group_member_info(filename):
        wb = load_workbook(filename=filename, read_only=False)
        sheet = wb.active
        items = []
        for index, r in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0:
                continue
            item = list(r)
            role_info = item[-1]
            if role_info is None:
                continue
            role_info_list = [_ for _ in role_info.split("\n") if _]
            group_id = [item[0], item[2], item[4]]
            group_id = [_ for _ in group_id if _ != "/"][0]
            for role_str in role_info_list:
                username, role = role_str.split(';')
                items.append([group_id, username, role])
        return items

    def create_group_member(self, group_id, username, role):
        menu_object = MenuModel.objects.filter(id=group_id).first()
        user_object = UserProfile.objects.filter(username=username).first()
        if not user_object:
            self.stdout.write(self.style.NOTICE(f'user {username} does not exist'))
            return False
        if not menu_object:
            self.stdout.write(self.style.NOTICE(f'group {group_id} does not exist'))
            return False
        member = Menu2Member.objects.filter(menu=menu_object, member=user_object).first()
        if member:
            if member.role != role:
                member.role = role
                member.save()
            return False

        _, created_status = Menu2Member.objects.get_or_create(
            menu=menu_object,
            member=user_object,
            role=role,
        )
        return created_status
