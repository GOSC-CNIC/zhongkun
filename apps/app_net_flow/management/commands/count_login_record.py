from openpyxl import load_workbook
import os
import sys

import random
import string
from django import setup
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError

# 将项目路径添加到系统搜寻路径当中，查找方式为从当前脚本开始，找到要调用的django项目的路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'django_site.settings')
setup()

from apps.app_users.models import UserProfile
from apps.app_net_flow.models import Menu2Member
from apps.app_net_flow.models import MenuModel


class Command(BaseCommand):
    help = """
    python manage.py count_login_record --file="/home/x.xlsx"
    Title format: ['通行证账号', '姓', '名']
    """

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', dest='filename', type=str, required=True,
            help='path of file',
        )
        parser.add_argument(
            '--start', dest='start', type=int, required=True,
            help='start timestamp',
        )

        parser.add_argument(
            # 当命令行有此参数时取值const, 否则取值default
            '--test', default=None, nargs='?', dest='test', const=True,
            help='test.',
        )

    def handle(self, *args, **options):
        is_test = options['test']
        filename = options['filename']
        start = options['start']

        if not Path(filename).exists():
            raise CommandError('not found file')

        user_info_list = self.parser_user_info(filename=filename)
        user_info_len = len(user_info_list)

        self.stdout.write(self.style.WARNING(f'excel_rows: {user_info_len};'))
        self.stdout.write(self.style.WARNING(f'start_timestamp: {start};'))

        if is_test:
            self.stdout.write(self.style.NOTICE('In mode Test'))
        active_user_count = 0
        for userinfo in user_info_list:
            username = userinfo[0]
            user_object = UserProfile.objects.filter(username=username).first()
            if user_object:
                last_active = user_object.last_active
                last_active_timestamp = last_active.timestamp()
                if last_active_timestamp >= start:
                    active_user_count += 1
        self.stdout.write(self.style.WARNING(f'active_user_count: {active_user_count};'))

    @staticmethod
    def parser_user_info(filename):
        wb = load_workbook(filename=filename, read_only=False)
        sheet = wb.active
        items = []
        for r in sheet.iter_rows(values_only=True):
            item = list(r)
            item = item[:3]
            if item[0] == '通行证账号':
                continue
            items.append(item)
        return items
